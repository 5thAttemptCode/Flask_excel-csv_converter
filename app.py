from flask import Flask, request, send_file, render_template
from openpyxl import load_workbook
import csv
import io
import os

app = Flask(__name__)

# Route to upload the Excel file
@app.route('/')
def upload_file():
    return render_template('index.html')

# Route to process the Excel file and convert it to CSV
@app.route('/convert', methods=['POST'])
def convert_excel_to_csv():
    # Check if the post request has the file part
    if 'file' not in request.files:
        return "No file part"
    
    file = request.files['file']

    # Check if the file is selected
    if file.filename == '':
        return "No selected file"
    
    # Extract the file name without extension
    original_filename = os.path.splitext(file.filename)[0]

    # Load the uploaded Excel file
    wb = load_workbook(file)
    sheet = wb.active

    # Prepare CSV data in-memory using io.StringIO
    output = io.StringIO()
    writer = csv.writer(output)
    
    for row in sheet.iter_rows(values_only=True):
        writer.writerow(list(row))
    
    output.seek(0)  # Go back to the start of the StringIO object
    
    # Set the download file name as the original file name with .csv extension
    csv_filename = f"{original_filename}.csv"
    
    # Send the CSV file as an attachment with the dynamic file name
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=csv_filename
    )

if __name__ == '__main__':
    app.run(debug=True)
